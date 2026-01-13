import os
from openpyxl import Workbook, load_workbook
import msvcrt

class empDB:
    def __init__(self, filename, recyclebin , headers):
        self.filename = filename
        self.recyclebin = recyclebin

        # Main database 
        if not os.path.isfile(self.filename):
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(headers)
            self.headers = headers
            self.wb.save(self.filename)
        else:
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
            self.headers = [cell.value for cell in self.ws[1]]

        # Recycle Database
        if not os.path.isfile(self.recyclebin):
            self.re_wb = Workbook()
            self.re_ws = self.re_wb.active
            self.re_ws.append(headers)
            self.re_wb.save(self.recyclebin)
        else:
            self.re_wb = load_workbook(self.recyclebin)
            self.re_ws = self.re_wb.active

    def add(self, empdata):
        for i in empdata:
            if i == "":
                print("Please provide enough data")
                print(self.headers)
                return
        
        self.ws.append(empdata)
        self.wb.save(self.filename)
        print("Data Added")

    def find(self, ID):
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] == ID:
                print(row)
                return
        print(f"The data for 'EMPID:{ID}' is not present in database")
        return

    def modify(self, ID):
        for row in self.ws.iter_rows(min_row=1,):
            if row[0].value == ID:
                print("Leave Empty to skip Modication")
                for i in range(1, len(self.headers)):
                    new_data = input(f"Enter new {headers[i]} : ")
                    if new_data != "":
                        row[i].value = new_data
                self.wb.save(self.filename)
                print("Data Modified")
                return
        print(f"Data for 'EMPId:{ID}' not present in database")
        return

    def recover(self, ID=None, recover_all = False):
        rows_to_recover = []

        for row_num, row in enumerate(self.re_ws.iter_rows(min_row=2), start=2):
            if(recover_all or row[0].value == ID or recover_all):
                rows_to_recover.append((row_num,[cell.value for cell in row]))
        
        if not rows_to_recover:
            print(f"No data to recover regarding 'EMPId:{ID}'")

        for _, data in rows_to_recover:
            self.ws.append(data)    
            print(f"Recovered: {data}")
            
        for row_num,_ in reversed(rows_to_recover):
            self.re_ws.delete_rows(row_num)
        
        self.wb.save(self.filename)
        self.re_wb.save(self.recyclebin)

    def clear_data(self):
        for row_num in range(self.re_ws.max_row, 1, -1):
            self.re_ws.delete_rows(row_num)
        self.re_wb.save(self.recyclebin)
        print("Recycle Bin Cleared")
        return


    def delete(self, ID):
        for row_num, row in enumerate(self.ws.iter_rows(min_row=2), start=2):
            if row[0].value == ID:
                temp_data = [cell.value for cell in row]
                print(temp_data)
                choice = input("Are you sure do you want to delete this data Y/N: ")
                if "y" == choice.lower():
                    self.ws.delete_rows(row_num)
                    print("Data Deleted")
                else:
                    print("Data not deleted")
                self.wb.save(self.filename)
                self.re_ws.append(temp_data)
                self.re_wb.save(self.recyclebin)
                return
        print(f"The data for 'EMPID:{ID}' is not present in database")
        return



if __name__ == "__main__":
    headers = ["EMPID","EmpName","Dept","Salary"]
    db = empDB(filename="Employee_databse.xlsx", recyclebin="RecycleDB.xlsx", headers=headers)

    while True:
        print("1. Add Employee")
        print("2. Find Employee")
        print("3. Modify Employee")
        print("4. Delete Employee")
        print("5. Reycle Bin")
        print("6. Exit")


        ch = msvcrt.getch().decode()
        match ch:
            case "1":
                data = dict.fromkeys(headers, "")
                for i in data.keys():
                    d = input(f'Enter {i} : ')
                    data[i] = d
                db.add(list(data.values()))
            case "2":
                id = input("Enter EmpID: ")
                db.find(id)
                
            case "3":
                id = input("Enter EmpID: ")
                db.modify(id)
    
            case "4":
                id = input("Enter Empid: ")
                db.delete(id)

            case "5":
                os.system('cls')
                print("1. Recover Data")
                print("2. Recover All")
                print("3. Clear Recycle Bin")
                re_ch = msvcrt.getch().decode()
                match re_ch:
                    case "1":
                        id = input("Enter EmpID: ")
                        db.recover(ID= id)
                    case "2":
                        db.recover(recover_all=True)
                    case "3":
                        db.clear_data()
            case "6":
                break
